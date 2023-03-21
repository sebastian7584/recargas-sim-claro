from selenium.common.exceptions import NoSuchElementException, NoSuchFrameException, NoSuchWindowException
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from subprocess import CREATE_NO_WINDOW
from openpyxl import load_workbook
import chromedriver_autoinstaller
from selenium import webdriver
from subprocess import Popen
from turtle import title
from tkinter import *
from tkinter import ttk
import pandas as pd
import numpy as np
import time


root = Tk()
root.title("")
root.geometry('500x500')
root.config(bg= '#fff')

link = "https://atiendo.claro.com.co/pretups/"
link2 = "https://atiendo.claro.com.co/pretups/c2sRechargeAction.do?method=c2sRechargeAuthorize&amp;moduleCode=C2STRF"

global user_key
global password_key
global pin
global tiempoOperacion

user_key = StringVar(root, "")
password_key = StringVar(root, "")
pin = StringVar(root, "")
tiempoOperacion = StringVar(root, "")
checkVariable = BooleanVar(root, False)
paqueteSelect = StringVar(root, "")




colorTeam = ('#E11419')



def openChrome(link):
    service = ChromeService('chromedriver')
    service.creationflags = CREATE_NO_WINDOW
    options =  webdriver.ChromeOptions()
    # options.add_argument('--headless')
    # options.add_experimental_option('excludeSwitches', ['enable-logging'])
    # options.add_argument('--log-level=3')
    browser = webdriver.Chrome(chrome_options= options)
    browser.get(link)
    # browser.set_window_position(0, 0)
    # browser.set_window_size(0, 0)
    browser.minimize_window()
    return browser

def excel():
    file = "lineas.xlsx"
    fileExcel = pd.read_excel(file)
    numbers = np.asarray(fileExcel)
    return numbers

def deleteCeldExcel(celd):
    file = "lineas.xlsx"
    workbook = load_workbook(file)
    sheet = workbook.active
    sheet['A'+str(celd)] = ""
    sheet['B'+str(celd)] = ""
    workbook.save(filename=file) 

def writeCeldExcel(celd,number,amount):
    file = "lineas.xlsx"
    workbook = load_workbook(file)
    sheet = workbook.active
    sheet['A'+str(celd)] = number
    sheet['B'+str(celd)] = amount
    workbook.save(filename=file)

def insert(by, str, text, browser):
    if by == "xpath": find = browser.find_element_by_xpath(str)
    elif by == "id": find = browser.find_element_by_id(str)
    elif by == "name": find = browser.find_element_by_name(str)
    else: find =None
    if find is not None:
        find.send_keys(text)

def click(by, str, browser):
    if by == "xpath": find = browser.find_element_by_xpath(str)
    elif by == "id": find = browser.find_element_by_id(str)
    elif by == "name": find = browser.find_element_by_name(str)
    else: find =None
    if find is not None:
        find.click()

def login(user_key, password_key, browser):
    archivo = pd.read_csv("data.csv")
    dataUser = np.asarray(archivo)
    user_key.set(dataUser[0][1])
    user_key2 = user_key.get() 
    password_key.set(dataUser[1][1]) 
    password_key2= password_key.get()
    

    insert("id", "loginID", user_key2, browser)
    insert("id", "password", password_key2, browser)
    time.sleep(0.5)
    click("name", "submit1", browser)
    time.sleep(0.5)
    browser.switch_to.frame("mainFrame")

def recharge(number,amount,pin, browser):
    groupId = "Paquetes de voz"
    archivo = pd.read_csv("data.csv")
    dataUser = np.asarray(archivo)
    pin.set(dataUser[2][1])
    pin2=pin.get()
    if checkVariable.get(): insert("name", "groupId", groupId, browser)
    if checkVariable.get(): insert("name", "subServiceType", paqueteSelect.get(), browser)
    insert("name", "subscriberMsisdn", number, browser)
    if checkVariable.get()== False: insert("name", "amount", amount, browser)
    insert("name", "pin", pin2, browser)
    click("name", "btnSubmit", browser)
    time.sleep(0.5)
    click("name", "btnSubmit", browser)
    time.sleep(0.5)
    click("name", "btnBack", browser)
    time.sleep(0.5)

def checklink(browser):
    browser.switch_to.frame("mainFrame")
    
def cargarLineas():
    root.withdraw()
    browser = openChrome(link)
    try:
        try:
            login(user_key, password_key, browser)
        except(NoSuchElementException, NoSuchFrameException):
            # browser.quit()
            # browser = openChrome(link)
            browser.get(link)
            try:
                login(user_key, password_key, browser)
            except (NoSuchElementException, NoSuchFrameException):
                pass
            time.sleep(1)
        time.sleep(0.5)
        numbers = excel()
        celd = 2
        errors = []
        for i in numbers:
            try:
                number = str(i[0])
                amount = str(i[1])
                recharge(number, amount, pin, browser)
                deleteCeldExcel(celd)
                celd+=1
                if tiempoOperacion.get== "1": tiempoRecarga = 1
                elif tiempoOperacion.get=="2": tiempoRecarga = 2
                elif tiempoOperacion.get=="3": tiempoRecarga = 3
                elif tiempoOperacion.get=="4": tiempoRecarga = 4
                elif tiempoOperacion.get=="5": tiempoRecarga = 5
                else: tiempoRecarga = 2
                time.sleep(tiempoRecarga)
            except(NoSuchElementException, NoSuchFrameException):
                errors.append([number,amount])
                # browser.quit()
                # browser = openChrome(link)
                browser.get(link)
                try:
                    login(user_key, password_key, browser)
                except (NoSuchElementException, NoSuchFrameException):
                    pass
                time.sleep(1)
                pass
        celdError = 2
        browser.quit()
        for error in errors:
            #writeCeldExcel(celdError,errors[error][0], errors[error][1])
            celdError+=1
        root.iconify()
        root.deiconify()
    except NoSuchWindowException:
        root.iconify()
        root.deiconify()

def abrirArchivo():
   p = Popen("openExcel.bat")
   stdout, stderr = p.communicate()
   listbox.insert(END,"Excel abierto, cerrar antes de proseguir")
   listbox.insert(END,"----------------------------------------")

def escribirDatos(win,title, str, x, y):
    data = Entry(win, textvariable=str,)
    data.delete("0",END)
    data.insert(0, str)
    data2= Label(win, text=title)
    data2.place(rely=y+0.05, relx= x)
    data2.config(font= ("Verdana", 12))
    data.place(relx=x+0.35, rely=y+0.05, relwidth=0.35, relheight=0.10)

def cargarDatos():
    archivo = pd.read_csv("data.csv")
    dataUser = np.asarray(archivo)
    user_key.set(dataUser[0][1]) 
    password_key.set(dataUser[1][1]) 
    pin.set(dataUser[2][1]) 
    tiempoOperacion.set(dataUser[3][1])

def guardarDatos():
     
    archivo = pd.read_csv("data.csv")
    archivo.loc[0,"data"] = user_key.get()
    archivo.loc[1,"data"] = password_key.get()
    archivo.loc[2,"data"] = pin.get()
    archivo.loc[3,"data"] = tiempoOperacion.get()
    archivo.to_csv("data.csv", index=False)
    print(user_key.get())

def configuraciones():
    confRoot = Toplevel(root)
    confRoot.title("Configuraciones")
    confRoot.geometry("300x200")
    escribirDatos(confRoot, "Usuario", user_key, 0.01, 0.01)
    escribirDatos(confRoot, "Clave", password_key, 0.01, 0.16)
    escribirDatos(confRoot, "Pin", pin, 0.01, 0.31)
    escribirDatos(confRoot, "Tiempo", tiempoOperacion, 0.01, 0.46)
    cargar = Button(confRoot, text="CARGAR", command= cargarDatos, bg= colorTeam, fg= 'white')
    cargar.place(relx=0.1,rely=0.75, relwidth=0.35, relheight=0.15)
    guardar = Button(confRoot, text="GUARDAR", command= guardarDatos, bg= colorTeam, fg= 'white')
    guardar.place(relx=0.5,rely=0.75, relwidth=0.35, relheight=0.15)

chromedriver_autoinstaller.install()
titulo = Label(root, text= "Team Comunicaciones")
titulo.pack(anchor= CENTER)
titulo.config(fg = colorTeam, bg= "white" ,font= ("Verdana", 24))


listbox = Listbox(root)
listbox.place(relx=0.04, rely=0.1, relwidth=0.6, relheight=0.6)

barra = Scrollbar(root, command=listbox.yview)
barra.place(relx=0.64, rely=0.1, relheight=0.6)

listbox.config(yscrollcommand=barra)

button1 = Button(root, text="RECARGAS WEB", command= cargarLineas, bg= colorTeam, fg= 'white')
button1.place(relx=0.05,rely=0.8, relwidth=0.20, relheight=0.1)

button2 = Button(root, text="ABRIR LISTA", command= abrirArchivo, bg= colorTeam, fg= 'white')
button2.place(relx=0.30,rely=0.8, relwidth=0.20, relheight=0.1)

imagen = PhotoImage(file ='logo.png')
lbImagen = Label(root, image= imagen, bd=0, fg="white").place(relx=0.65,rely=0.7)

marca = Label(root, text= "Sebastian Moncada Cel:324-221-0852 ")
marca.pack()
marca.config(fg = "black", bg= "white" ,font= ("Verdana", 8))
marca.place(relx=0.01,rely=0.95)

imgConf =PhotoImage(file='conf.png')
conf = Button(text="test", width=20, height=20, image=imgConf,justify="right", bd=0, fg="white", bg="white", command= configuraciones)
conf.place(relx=0.05, rely=0.02)

checkbox = Checkbutton(text= "PAQUETES", variable=checkVariable, command=lambda: paquetes(), bg='white')
checkbox.place(relx=0.75, rely=0.15)

frame = Frame(root, bg='white',height=200, width=125)
listaPaquetes2 = [
    "Paquetes Todo Incluido",
    "Paquetes de datos",
    "Reventa",
    "Aplicaciones y Redes",
    "Paquetes de Voz",
    "Paquetes de LDI",
    "Paquetes GAMERS",
    ]
listaPaquetes = [
    "VZ - Paq 50 Min – 1 Dia - $1,000",
    "VZ - Paq 120 Min - 1 Dia - $ 2,000",
    "VZ- Paq 60 Min - 2 Dias - $3,000",
    "VZ- Paq 100 Min TAT - $5,000",
    "VZ- Paq 200 Min TAT - $9,900",
    
]
paquetes = ttk.Combobox(frame, values=listaPaquetes, textvariable= paqueteSelect)
paquetes.place(relx=0.01, rely=0.01, width =125, height=25)



def paquetes():
    if frame.winfo_ismapped():
        frame.place_forget()
    else:
        frame.place(relx=0.70, rely=0.2)

        




def app(root):
    root.mainloop()
   
   

app(root)
